"""Export sessions to an .xlsx file (openpyxl)."""
from __future__ import annotations

import datetime
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

_HEADERS = ["Date", "Day", "Start", "End", "Room", "Role",
            "Primary Author", "Session Code", "Title", "Event"]

_COL_WIDTHS = {
    "Date": 13, "Day": 12, "Start": 7, "End": 7, "Room": 22,
    "Role": 12, "Primary Author": 14, "Session Code": 13,
    "Title": 60, "Event": 18,
}


def _extract_year(event_name: str) -> int | None:
    m = re.search(r"\b(20\d{2})\b", event_name or "")
    return int(m.group(1)) if m else None


def _parse_date(date_str: str, year: int) -> tuple[int, int, int] | None:
    m = re.search(r"([A-Za-z]+)\.?\s+(\d{1,2})", date_str or "")
    if not m:
        return None
    key = m.group(1).lower()[:3]
    mo = _MONTH_MAP.get(key)
    return (year, mo, int(m.group(2))) if mo else None


def _parse_time(time_str: str) -> tuple[tuple[int, int], tuple[int, int]] | None:
    m = re.match(r"(\d{1,2}):(\d{2})\s*[-–]\s*(\d{1,2}):(\d{2})", (time_str or "").strip())
    return ((int(m.group(1)), int(m.group(2))), (int(m.group(3)), int(m.group(4)))) if m else None


def _sort_key(s: dict, year: int | None) -> tuple:
    ymd = _parse_date(s.get("date", ""), year) if year else None
    times = _parse_time(s.get("time", ""))
    return (
        ymd[1] if ymd else 99,
        ymd[2] if ymd else 99,
        times[0][0] if times else 99,
        times[0][1] if times else 99,
    )


def export_xlsx(
    sessions: list[dict[str, Any]],
    event_timezone: str,
    output_path: str | Path,
    event_name: str = "",
    event_location: str = "",
    query: str = "",
) -> None:
    year = _extract_year(event_name)
    sorted_sessions = sorted(sessions, key=lambda s: _sort_key(s, year))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sessions"

    arial = "Arial"
    header_font = Font(name=arial, bold=True, size=10)
    data_font = Font(name=arial, size=10)
    meta_font = Font(name=arial, size=10, italic=True, color="595959")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")

    # Meta rows
    meta_rows = [
        ("Event", event_name),
        ("Location", event_location),
        ("Timezone", event_timezone),
    ]
    if query:
        meta_rows.append(("Query", query))

    for label, value in meta_rows:
        ws.append([f"{label}: {value}"])
        ws.cell(ws.max_row, 1).font = meta_font

    # Header row
    ws.append(_HEADERS)
    header_row = ws.max_row
    for col, _ in enumerate(_HEADERS, 1):
        cell = ws.cell(header_row, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = top_align

    ws.freeze_panes = f"A{header_row + 1}"

    # Data rows
    for s in sorted_sessions:
        ymd = _parse_date(s.get("date", ""), year) if year else None
        times = _parse_time(s.get("time", ""))

        date_val = datetime.date(*ymd) if ymd else None
        day_name = date_val.strftime("%A") if date_val else ""
        if times:
            (sh, sm), (eh, em) = times
            start_val = datetime.time(sh, sm)
            end_val = datetime.time(eh, em)
        else:
            start_val = end_val = None

        role_raw = (s.get("role") or "unknown").lower()
        role_display = {"speaker": "Speaker", "chair": "Chair", "discussant": "Discussant"}.get(
            role_raw, role_raw.capitalize()
        )
        primary = "Y" if s.get("is_primary_author") else "N"
        title = s.get("talk_title") or s.get("session_title") or ""

        row = [
            date_val,
            day_name,
            start_val,
            end_val,
            s.get("room") or "",
            role_display,
            primary,
            s.get("session_code") or "",
            title,
            event_name,
        ]
        ws.append(row)
        data_row = ws.max_row

        for col in range(1, len(_HEADERS) + 1):
            cell = ws.cell(data_row, col)
            cell.font = data_font
            cell.alignment = wrap_align

        # Date format
        if date_val is not None:
            ws.cell(data_row, 1).number_format = "YYYY-MM-DD"
        # Time format
        for col in (3, 4):
            if ws.cell(data_row, col).value is not None:
                ws.cell(data_row, col).number_format = "HH:MM"

    # Column widths
    for col, header in enumerate(_HEADERS, 1):
        ws.column_dimensions[get_column_letter(col)].width = _COL_WIDTHS.get(header, 15)

    wb.save(output_path)
